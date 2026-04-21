import io
import re
from datetime import date, datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

DEFAULT_FILE = Path(__file__).resolve().parent / "Análises CIE 2025 e 2026.xlsx"
ALLOWED_ISOTOPES = ["HO", "C", "HCO", "NC", "N", "CO"]
ISOTOPE_SET_TO_LABEL = {
    frozenset({"H", "O"}): "HO",
    frozenset({"C"}): "C",
    frozenset({"H", "C", "O"}): "HCO",
    frozenset({"N", "C"}): "NC",
    frozenset({"N"}): "N",
    frozenset({"C", "O"}): "CO",
}
YELLOW_COLORS = {"FFFFFF00", "FFFFF2CC"}


def normalize_color(color_value) -> str:
    if color_value is None:
        return ""
    color_text = str(color_value).upper().replace("#", "")
    return color_text[-8:]


def canonical_isotope_label(letters: set[str]) -> str:
    order = ["N", "C", "H", "O"]
    return "".join(letter for letter in order if letter in letters)


def is_completed_row(cells) -> bool:
    for cell in cells:
        fill = cell.fill
        if fill is None or fill.patternType != "solid":
            continue
        color = normalize_color(fill.fgColor.rgb if fill.fgColor.type == "rgb" else None)
        if color in YELLOW_COLORS:
            return True
    return False


def parse_quantity(value) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    match = re.search(r"\d+(?:[.,]\d+)?", str(value))
    if not match:
        return 0.0
    return float(match.group(0).replace(",", "."))


def infer_year(raw_date, current_year: int | None, solicitation: str | None) -> int | None:
    if isinstance(raw_date, (datetime, date)):
        return raw_date.year

    if isinstance(raw_date, (int, float)) and float(raw_date).is_integer():
        year = int(raw_date)
        if 2000 <= year <= 2100:
            return year

    if isinstance(raw_date, str) and raw_date.strip():
        parsed = pd.to_datetime(raw_date, dayfirst=True, errors="coerce")
        if pd.notna(parsed):
            return int(parsed.year)

        year_match = re.search(r"(20\d{2})", raw_date)
        if year_match:
            return int(year_match.group(1))

    if solicitation:
        code_match = re.search(r"(\d{2})", str(solicitation))
        if code_match:
            return 2000 + int(code_match.group(1))

    return current_year


def normalize_isotope(code: str | None) -> str | None:
    if code is None:
        return None

    letters = {char for char in str(code).upper() if char in {"H", "N", "C", "O"}}
    if not letters:
        return None

    return ISOTOPE_SET_TO_LABEL.get(frozenset(letters), canonical_isotope_label(letters))


def isotope_multiplier(isotope: str | None) -> int:
    if not isotope:
        return 0
    if isotope == "HO":
        return 1
    return len(isotope)


def read_workbook(source_bytes: bytes | None, file_path: str):
    if source_bytes is not None:
        return load_workbook(io.BytesIO(source_bytes), data_only=True)

    workbook_path = Path(file_path).expanduser()
    if not workbook_path.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {workbook_path}")

    return load_workbook(workbook_path, data_only=True)


def parse_file_paths(raw_text: str) -> list[str]:
    return [line.strip() for line in raw_text.splitlines() if line.strip()]


def build_dataframe(workbook, source_name: str = "") -> pd.DataFrame:
    sheet = workbook[workbook.sheetnames[0]]
    current_year = None
    records = []

    for row in sheet.iter_rows(min_row=2, max_col=12):
        first_value = row[0].value

        if isinstance(first_value, (int, float)) and float(first_value).is_integer():
            year_marker = int(first_value)
            if 2000 <= year_marker <= 2100:
                current_year = year_marker
                continue

        solicitation = row[1].value
        raw_isotope = row[5].value
        if not solicitation or not raw_isotope:
            continue

        year = infer_year(row[0].value, current_year, solicitation)
        isotope = normalize_isotope(raw_isotope)
        quantity = parse_quantity(row[4].value)

        if year is None or isotope is None or quantity <= 0:
            continue

        records.append(
            {
                "ano": int(year),
                "arquivo": source_name,
                "solicitacao": str(solicitation).strip(),
                "status": str(row[2].value).strip() if row[2].value else "",
                "qte": quantity,
                "isotopo_original": str(raw_isotope).strip().upper(),
                "isotopo_normalizado": isotope,
                "concluida": is_completed_row(row[:11]),
            }
        )

    dataframe = pd.DataFrame(records)
    if dataframe.empty:
        return dataframe

    dataframe["multiplicador"] = dataframe["isotopo_normalizado"].apply(isotope_multiplier)
    dataframe["total"] = dataframe["qte"] * dataframe["multiplicador"]
    return dataframe


def load_dataframes(uploaded_files, file_paths: list[str]) -> pd.DataFrame:
    dataframes = []

    if uploaded_files:
        for uploaded_file in uploaded_files:
            workbook = read_workbook(uploaded_file.getvalue(), "")
            dataframes.append(build_dataframe(workbook, source_name=uploaded_file.name))
    else:
        for file_path in file_paths:
            workbook = read_workbook(None, file_path)
            dataframes.append(build_dataframe(workbook, source_name=Path(file_path).name))

    valid_frames = [dataframe for dataframe in dataframes if not dataframe.empty]
    if not valid_frames:
        return pd.DataFrame()

    return pd.concat(valid_frames, ignore_index=True)


def summarize_results(
    dataframe: pd.DataFrame,
    selected_years: list[int],
    selected_isotopes: list[str],
    completed_only: bool,
) -> pd.DataFrame:
    filtered = dataframe[dataframe["ano"].isin(selected_years)].copy()
    filtered = filtered[filtered["isotopo_normalizado"].isin(selected_isotopes)]

    if completed_only:
        filtered = filtered[filtered["concluida"]]

    if filtered.empty:
        return pd.DataFrame(columns=["ano", "isotopo", "solicitacoes", "analises", "total"])

    summary = (
        filtered.groupby(["ano", "isotopo_normalizado"], as_index=False)
        .agg(
            solicitacoes=("solicitacao", "count"),
            analises=("qte", "sum"),
            total=("total", "sum"),
        )
        .rename(columns={"isotopo_normalizado": "isotopo"})
    )

    summary["isotopo"] = pd.Categorical(summary["isotopo"], categories=ALLOWED_ISOTOPES, ordered=True)
    summary = summary.sort_values(["ano", "isotopo"]).reset_index(drop=True)
    return summary


def format_number(value: float) -> str:
    if pd.isna(value):
        return "0"
    if abs(value - round(value)) < 1e-9:
        return f"{int(round(value)):,}".replace(",", ".")
    return f"{value:,.2f}".replace(",", "_").replace(".", ",").replace("_", ".")


def build_markdown_report(summary: pd.DataFrame, completed_only: bool) -> str:
    status_text = "somente solicitações concluídas" if completed_only else "todas as solicitações"
    lines = ["## Mini-relatório", f"Filtro aplicado: **{status_text}**.", ""]

    if summary.empty:
        lines.append("Nenhum registro foi encontrado com os filtros escolhidos.")
        return "\n".join(lines)

    for record in summary.itertuples(index=False):
        solicitacao_label = "solicitação" if record.solicitacoes == 1 else "solicitações"
        analise_label = "análise" if record.analises == 1 else "análises"
        lines.append(
            f"- Em **{record.ano}**, para o isótopo **{record.isotopo}**, "
            f"foram registradas **{format_number(record.analises)} {analise_label}** "
            f"em **{record.solicitacoes} {solicitacao_label}**, "
            f"com **total de {format_number(record.total)}**."
        )

    lines.extend(
        [
            "",
            f"**Total geral de análises:** {format_number(summary['analises'].sum())}",
            f"**Total geral consolidado:** {format_number(summary['total'].sum())}",
            f"**Total geral de solicitações:** {format_number(summary['solicitacoes'].sum())}",
        ]
    )
    return "\n".join(lines)


def main():
    import streamlit as st

    st.set_page_config(page_title="Relatório CIE", layout="wide")
    st.title("Relatório de Análises CIE")
    st.write("Use a barra lateral para filtrar os anos, os isótopos e o tipo de contagem.")

    with st.sidebar:
        st.header("Filtros")
        uploaded_files = st.file_uploader(
            "Arquivo(s) Excel (.xlsx)",
            type=["xlsx"],
            accept_multiple_files=True,
        )
        raw_file_paths = st.text_area(
            "Caminho(s) do(s) arquivo(s), um por linha",
            value=str(DEFAULT_FILE),
            help="Se você fizer upload de arquivo(s), eles terão prioridade sobre os caminhos digitados.",
        )
        show_completed_only = st.radio(
            "Quais solicitações deseja contabilizar?",
            options=["Todas", "Somente concluídas"],
            index=0,
        )
        show_chart = st.checkbox("Gerar gráfico de barras", value=True)

    try:
        file_paths = parse_file_paths(raw_file_paths)
        dataframe = load_dataframes(uploaded_files, file_paths)
    except Exception as exc:
        st.error(f"Não foi possível ler o arquivo: {exc}")
        st.stop()

    if dataframe.empty:
        st.warning("Nenhum dado válido foi encontrado na planilha.")
        st.stop()

    available_years = sorted(dataframe["ano"].dropna().unique().tolist())

    with st.sidebar:
        selected_years = st.multiselect("Ano(s)", options=available_years, default=available_years)
        selected_isotopes = st.multiselect("Isótopo(s)", options=ALLOWED_ISOTOPES, default=ALLOWED_ISOTOPES)

    if not selected_years or not selected_isotopes:
        st.info("Selecione pelo menos um ano e um isótopo para gerar o relatório.")
        st.stop()

    completed_only = show_completed_only == "Somente concluídas"
    summary = summarize_results(dataframe, selected_years, selected_isotopes, completed_only)
    report_markdown = build_markdown_report(summary, completed_only)

    ignored_isotopes = sorted(
        value for value in dataframe["isotopo_normalizado"].dropna().unique().tolist() if value not in ALLOWED_ISOTOPES
    )
    if ignored_isotopes:
        st.info(
            "Existem combinações fora da lista principal de filtros "
            f"({', '.join(ignored_isotopes)}). Elas ficam disponíveis no tratamento interno, "
            "mas não aparecem na seleção padrão."
        )

    metric_col_1, metric_col_2, metric_col_3 = st.columns(3)
    metric_col_1.metric("Solicitações", format_number(summary["solicitacoes"].sum() if not summary.empty else 0))
    metric_col_2.metric("Análises", format_number(summary["analises"].sum() if not summary.empty else 0))
    metric_col_3.metric("Total", format_number(summary["total"].sum() if not summary.empty else 0))

    file_count = dataframe["arquivo"].nunique() if "arquivo" in dataframe.columns else 0
    st.caption(f"Base consolidada com {file_count} arquivo(s).")

    st.subheader("Tabela-resumo")
    if summary.empty:
        st.warning("Nenhum resultado encontrado para os filtros selecionados.")
    else:
        display_df = summary.copy()
        display_df["analises"] = display_df["analises"].map(format_number)
        display_df["total"] = display_df["total"].map(format_number)
        st.dataframe(display_df, use_container_width=True, hide_index=True)

    st.markdown(report_markdown)

    if show_chart and not summary.empty:
        st.subheader("Gráfico de barras")
        chart_data = (
            summary.assign(isotopo=summary["isotopo"].astype(str))
            .pivot(index="ano", columns="isotopo", values="total")
            .fillna(0)
        )
        st.bar_chart(chart_data)


if __name__ == "__main__":
    main()
